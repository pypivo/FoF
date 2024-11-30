import os
from openpyxl import Workbook, load_workbook

import default as d


def check_or_create_default_coefficients_xlsx(file_path, required_columns, data_dict):
    """
    Проверяет, существует ли xlsx файл и содержит ли он требуемые столбцы.
    Если файл не существует или столбцы отсутствуют, пересоздаёт файл 
    с данными из словаря.


    :param file_path: Путь к xlsx файлу.
    :param required_columns: Список обязательных названий столбцов.
    :param data_dict: Словарь с данными для записи в файл, где ключ — название столбца, а значение — данныe.
    """
    if os.path.exists(file_path):
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            # Получение списка существующих столбцов из первой строки
            existing_columns = [cell.value for cell in sheet[1]]
            # Проверка, есть ли все требуемые столбцы
            if all(column in existing_columns for column in required_columns):
                print("Файл существует и содержит все необходимые столбцы.")
                return
        except Exception as e:
            print(f"Ошибка при чтении файла: {e}")
    else:
        print("Файл не найден. Создаём новый...")

    # Создание нового файла
    workbook = Workbook()
    sheet = workbook.active

    # Запись названий столбцов
    column_names = list(data_dict.keys())
    for col_index, column_name in enumerate(column_names, start=1):
        sheet.cell(row=1, column=col_index, value=column_name)

    # Запись данных в столбцы
    for row_index in range(2, 3):
        for col_index, column_name in enumerate(column_names, start=1):
            value = data_dict[column_name]
            sheet.cell(row=row_index, column=col_index, value=value)

    workbook.save(file_path)
    print(f"Файл {file_path} создан/обновлён.")


if __name__ == '__main__':
    required_columns = list(d.default_coefficients.keys())
    data_dict = d.default_coefficients
    check_or_create_default_coefficients_xlsx("src/coefficients/coefficients.xlsx", required_columns, data_dict)
